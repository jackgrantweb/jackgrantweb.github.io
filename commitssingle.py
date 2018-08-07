import csv
import math
import time
import openpyxl

dest = "C://Users/JG31348/Documents/DLmetric.xlsx"
wb = openpyxl.load_workbook(filename = dest)
ws=wb.active

for i in range(111,168):
    sum=0
    count=0
    sum1=0
    sum2=0
    csvfile="C://Users/JG31348/Documents/commitsv22/commits"+str(i)+".csv"
    try:
        with open(csvfile,newline='') as csvfile1:
            spamreader=csv.reader(csvfile1,delimiter=",",quotechar='|')
            for row in spamreader:
                sum+=.1+((100+((int(row[0])-1)))/100)*(int(row[3])/1000)
                sum1+=int(row[3])
                sum2+=1
    except (FileNotFoundError):
        do=False
    print(str(i)+": "+str(sum)+": "+str(sum1)+": "+str(sum2))
    ws.cell(row=i-109,column=100).value=sum
    ws.cell(row=i-109,column=101).value=sum1
wb.save(dest)


#mymetricanalysis()
